"""
Joint-track overview tables -- the full node-level data appendix backing
Table 14 (the thesis's collapsed 7-scenario S1/S2/S2.1/S3/S4/S4.1/S5 table,
see docs/decisions.md), one sheet per treatment (K-plane /
Y-plane) since the two are never merged (see docs/decisions.md --
K/Y is an orthogonal axis, not folded into the scenario dimension).

Reproduces, by script rather than by hand, the table the author was building
directly in Excel from final_results_joint_summary.csv (conditional
green/pass vs orange/fail fill by D>=1) -- written to a SEPARATE new
workbook so it doesn't collide with that live, hand-edited file.

Source: results/real_campaign/final_results_joint_summary.csv (already
renamed to the D_S1-K/D_S1-Y/... convention, see that renaming session).
"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

POSTPRO_DIR = Path(__file__).resolve().parent
PROJECT = POSTPRO_DIR.parent   # repo root
RESULTS_DIR = PROJECT / "results"
OUT_PATH = RESULTS_DIR / "joint_results_overview_tables.xlsx"

SCENARIOS = ["S1", "S2", "S2.1", "S3", "S4", "S4.1", "S5"]
D_FAILURE = 1.0

PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PASS_FONT = Font(color="006100")
FAIL_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
FAIL_FONT = Font(color="7F6000")
HEADER_FONT = Font(bold=True)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_data():
    return pd.read_csv(RESULTS_DIR / "final_results_joint_summary.csv").set_index("node")


def write_sheet(wb, df, treatment, title):
    ws = wb.create_sheet(title)
    headers = ["node", "family", "Splash Zone"] + [f"D_{s}-{treatment}" for s in SCENARIOS]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for r, (node, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=r, column=1, value=int(node)).border = BORDER
        ws.cell(row=r, column=2, value=row["family"]).border = BORDER
        ws.cell(row=r, column=3, value="Yes" if row["in_splash_zone"] else "No").border = BORDER
        for c, s in enumerate(SCENARIOS, start=4):
            col = f"D_{s}-{treatment}"
            val = row[col]
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if pd.isna(val):
                cell.value = None
                continue
            cell.value = round(float(val), 3)
            cell.number_format = "0.000"
            if val >= D_FAILURE:
                cell.fill = FAIL_FILL
                cell.font = FAIL_FONT
            else:
                cell.fill = PASS_FILL
                cell.font = PASS_FONT

    for c, h in enumerate(headers, start=1):
        width = max(len(h) + 2, 10)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"


def build():
    df = load_data()
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, df, "K", "K-plane overview")
    write_sheet(wb, df, "Y", "Y-plane overview")
    wb.save(OUT_PATH)
    return OUT_PATH


if __name__ == "__main__":
    out_path = build()
    print(f"Wrote {out_path}")
