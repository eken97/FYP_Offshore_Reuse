"""
Reuse-track report figures -- the "acceptance criteria" Results section.
Reads results/real_campaign/reuse_classification.csv and
reuse_classification_bays.csv (produced by stage4_reuse_classification.py,
see docs/decisions.md) and produces:

- C1: L0->L1->L2->L3->L4 classification cascade (all 112 members + 4 bays,
  both retrofit verdicts).
- C2: bay-level L1 margin dumbbell -- worst joint D under Retrofit A vs
  Retrofit B, against the MARGIN_L1_JOINT_D25 threshold (0.25, NOT the D=1
  failure line other joint-track figures use -- a different, stricter
  threshold specific to the structural-reuse margin).

Follows the same final_candidates/final convention as member/joint tracks
(see docs/decisions.md).
"""
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from matplotlib.lines import Line2D

import fatigue_style as fs
import sd_geometry as sdg

NOT_IN_SCOPE_COLOR = "#b8b6ad"  # matches jacket_face_plot.py's NOT_ASSESSABLE_COLOR

POSTPRO_DIR = Path(__file__).resolve().parent
PROJECT = POSTPRO_DIR.parent   # repo root
RESULTS_DIR = PROJECT / "results"
CARBON_WORKBOOK = RESULTS_DIR / "embodied_carbon_inputs.xlsx"
BAY_GROUPING_CSV = PROJECT / "results" / "bay_grouping.csv"
OUT_DIR = PROJECT / "figures" / "reuse_track" / "final_candidates"

MARGIN_L1_JOINT_D25 = 0.25  # must match stage4_reuse_classification.py


def load():
    """Member-level results are scoped to the 80 "big" legs/braces that make
    up the 4 structural bays (bay_grouping.csv, entity_type=='member') --
    same scope decision already used for the embodied-carbon workbook's
    reusable-length cutoff and the L1 bay analysis itself. The excluded 32
    are short transition/stub members (foundation-adjacent legs + a few
    submerged stub braces near the pile transition) that are never candidates
    for reuse and go straight to recycling regardless of L0-L3 outcome --
    not interesting to show in the member-level reuse figures/tables.
    Confirmed the 32 splash-zone members are untouched by this filter (all
    32 sit inside the 4 bays), so the L3/Downgraded count is unaffected.
    """
    members_all = pd.read_csv(RESULTS_DIR / "reuse_classification.csv")
    bays = pd.read_csv(RESULTS_DIR / "reuse_classification_bays.csv")
    bay_grouping = pd.read_csv(BAY_GROUPING_CSV)
    big_ids = set(bay_grouping.loc[bay_grouping["entity_type"] == "member", "entity_id"])
    members = members_all[members_all["member_id"].isin(big_ids)].copy()
    n_excluded = len(members_all) - len(members)
    print(f"Scoped to {len(members)}/{len(members_all)} members "
          f"({n_excluded} small/foundation-transition members excluded)")
    return members, bays


# ---------------------------------------------------------------------------
# C1 -- classification cascade (funnel). One retrofit at a time is a single
# path L0 -> L1 -> L2 -> L3 -> L4; A and B give identical counts today (see
# memory), so this draws BOTH retrofit labels on the same bars rather than
# duplicating the whole figure -- the counts coincide, but the two arrows
# into L1 are drawn separately so a future divergence would be visible.
# ---------------------------------------------------------------------------

def fig_C1_cascade(members, bays):
    n_total = len(members)
    n_l0_pass = int(members["l0_pass"].sum())
    n_l1_a = int(bays["structural_reuse_pass_A"].sum())  # bays, not members
    n_l1_b = int(bays["structural_reuse_pass_B"].sum())
    n_l2 = int((members["reuse_category_A"] == "L2 - Component reuse").sum())
    n_l3 = int((members["reuse_category_A"] == "L3 - Downgraded reuse").sum())
    n_l4 = int((members["reuse_category_A"] == "L4 - Recycle").sum()) if \
        (members["reuse_category_A"] == "L4 - Recycle").any() else 0

    # Categorical identity colours (matches the blue/orange used everywhere
    # else in the report, e.g. Retrofit A/B dumbbells) -- NOT the reserved
    # status colours (green/yellow/red), which mean something different
    # (life vs the 25yr design threshold) elsewhere in the report and would
    # clash if reused here for a different meaning. L0/L1/L4 are empty or a
    # pure gate in this dataset, so they stay a neutral grey; L2 and L3 are
    # the two real outcomes and get blue/orange respectively.
    stages = [
        ("L0\nAdmissible", n_l0_pass, n_total, fs.BASELINE),
        ("L1\nStructural reuse\n(bays)", n_l1_a, len(bays), fs.BASELINE),
        ("L2\nComponent reuse", n_l2, n_total, fs.CAT_BLUE),
        ("L3\nDowngraded reuse", n_l3, n_total, fs.CAT_ORANGE),
        ("L4\nRecycle", n_l4, n_total, fs.BASELINE),
    ]

    fig, ax = plt.subplots(figsize=fs.usable_figsize(width_frac=1.0, aspect=0.42))
    x = np.arange(len(stages))
    heights = [s[1] for s in stages]
    denom = [s[2] for s in stages]
    colors = [s[3] for s in stages]
    labels = [s[0] for s in stages]

    bars = ax.bar(x, heights, color=colors, width=0.55, zorder=3,
                   edgecolor=fs.MARKER_EDGE, linewidth=fs.MARKER_EDGE_WIDTH)
    for xi, h, d in zip(x, heights, denom):
        pct = 100 * h / d
        ax.text(xi, h + max(heights) * 0.03, f"{h}/{d}\n({pct:.0f}%)",
                 ha="center", va="bottom", fontsize=7.5, color=fs.INK_PRIMARY)

    # L1's bar shows bays; annotate that its denominator is bays not members
    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=7.5, color=fs.INK_PRIMARY)
    ax.set_ylabel("Count")
    ax.set_ylim(0, n_total * 1.22)
    ax.spines[["top", "right"]].set_visible(False)
    ax.spines["left"].set_color(fs.BASELINE)
    ax.spines["bottom"].set_color(fs.BASELINE)
    ax.tick_params(colors=fs.INK_SECONDARY)
    ax.grid(axis="y", color=fs.GRIDLINE, linewidth=0.6, zorder=0)
    ax.set_axisbelow(True)

    # Note: L1 is counted per bay (n=4), L0/L2/L3/L4 per member (n=80); both
    # Retrofit A and B give the same L1 result (0/4). No in-image caption --
    # explained in the surrounding report text instead, per project convention.
    fig.tight_layout()
    return fs.save_fig(fig, OUT_DIR, "C1_reuse_classification_cascade")


# ---------------------------------------------------------------------------
# C2 -- bay-level L1 margin dumbbell. Worst joint D_A_K / D_B_K per bay
# against MARGIN_L1_JOINT_D25 (0.25) -- a stricter, different threshold than
# the D=1 Miner's-rule failure line other joint dumbbells use.
# ---------------------------------------------------------------------------

def fig_C2_bay_margin_dumbbell(bays):
    labels = [f"Bay {b}" for b in bays["bay_id"]]
    left_vals = bays["worst_joint_D_A_K"].values   # Retrofit A
    right_vals = bays["worst_joint_D_B_K"].values  # Retrofit B

    order = np.argsort(left_vals)[::-1]
    labels = [labels[i] for i in order]
    left_vals = left_vals[order]
    right_vals = right_vals[order]

    fig, ax = plt.subplots(figsize=fs.usable_figsize(width_frac=0.62, aspect=0.85))
    y = np.arange(len(labels))[::-1]
    for yi, lv, rv in zip(y, left_vals, right_vals):
        ax.plot([lv, rv], [yi, yi], color=fs.GRIDLINE, linewidth=1.2, zorder=1)
    ax.scatter(left_vals, y, color=fs.CAT_BLUE, s=48, zorder=3,
               label="Retrofit A -- worst joint D", edgecolor=fs.MARKER_EDGE,
               linewidth=fs.MARKER_EDGE_WIDTH)
    ax.scatter(right_vals, y, color=fs.CAT_ORANGE, s=48, zorder=3,
               label="Retrofit B -- worst joint D", edgecolor=fs.MARKER_EDGE,
               linewidth=fs.MARKER_EDGE_WIDTH)

    ax.axvline(MARGIN_L1_JOINT_D25, color=fs.STATUS_CRITICAL, linewidth=1.1,
               linestyle=":", zorder=2)
    ax.text(MARGIN_L1_JOINT_D25, len(labels) - 0.05,
            f"  L1 margin (D={MARGIN_L1_JOINT_D25})", fontsize=7,
            color=fs.STATUS_CRITICAL, va="top", ha="left")

    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=8.5)
    ax.set_ylim(-0.8, len(labels) - 0.2)
    ax.set_xscale("log")
    fs.add_log_gridlines(ax, axis="x", subs=range(2, 10))
    ax.set_xlabel("Worst joint D at year 25 (K consideration, log scale)")
    ax.tick_params(colors=fs.INK_SECONDARY)
    ax.spines[["top", "right"]].set_visible(False)
    ax.spines["left"].set_color(fs.BASELINE)
    ax.set_title("Bay-level structural-reuse (L1) margin check", loc="left",
                 fontsize=8.5, color=fs.INK_PRIMARY, fontweight="bold", pad=6)

    fig.tight_layout(rect=[0, 0.12, 1, 1])
    handles, legend_labels = ax.get_legend_handles_labels()
    fig.legend(handles, legend_labels, loc="lower center", ncol=2, frameon=False,
               fontsize=7.5, bbox_to_anchor=(0.5, 0.0))
    return fs.save_fig(fig, OUT_DIR, "C2_bay_l1_margin_dumbbell")


# ---------------------------------------------------------------------------
# C3 -- static isometric jacket map, coloured by reuse category. Print
# equivalent of the interactive 3D viewer (same orthographic, spherical-basis
# camera angle: azimuth=-0.65, elevation=0.28), for the thesis PDF where the
# interactive version obviously can't go.
# ---------------------------------------------------------------------------

def _iso_basis(azimuth, elevation):
    caz, saz = np.cos(azimuth), np.sin(azimuth)
    cel, sel = np.cos(elevation), np.sin(elevation)
    forward = np.array([caz * cel, saz * cel, sel])
    world_up = np.array([0.0, 0.0, 1.0])
    right = np.cross(forward, world_up)
    right /= np.linalg.norm(right)
    up = np.cross(right, forward)
    up /= np.linalg.norm(up)
    return right, up, forward


def _iso_project(points, center, right, up, forward):
    rel = points - center
    sx = rel @ right
    sy = rel @ up
    depth = rel @ forward
    return sx, sy, depth


def fig_C3_jacket_reuse_isometric(members, azimuth=-0.65, elevation=0.28):
    """members: the SCOPED (80-row) dataframe, used only to look up category
    by member_id -- geometry (incl. the 32 excluded members, drawn muted) is
    read fresh from SubDyn.dat via sd_geometry, same source as every other
    jacket map in the report."""
    model = sdg.read_subdyn_model(sdg.DEFAULT_SD_PATH)
    joints = model["joints"]  # id -> (x, y, z)
    cat_by_id = members.set_index("member_id")["reuse_category_A"].to_dict()

    all_xyz = np.array(list(joints.values()))
    center = (all_xyz.min(axis=0) + all_xyz.max(axis=0)) / 2
    right, up, forward = _iso_basis(azimuth, elevation)

    fs.apply_style()
    fig, ax = plt.subplots(figsize=fs.usable_figsize(width_frac=0.55, aspect=1.55))

    # reference rings for elevation context (MSL + splash zone bounds)
    plan_radius = np.max(np.abs(all_xyz[:, :2] - center[:2])) * 1.35
    theta = np.linspace(0, 2 * np.pi, 64)

    def draw_ring(z_ref, linestyle):
        ring = np.stack([
            center[0] + plan_radius * np.cos(theta),
            center[1] + plan_radius * np.sin(theta),
            np.full_like(theta, z_ref),
        ], axis=1)
        sx, sy, _ = _iso_project(ring, center, right, up, forward)
        ax.plot(sx, sy, color=fs.BASELINE, linewidth=0.7, linestyle=linestyle, zorder=1)

    draw_ring(0.0, (0, (2, 2)))
    draw_ring(sdg.ZONE_SPLASH_ZMIN, ":")
    draw_ring(sdg.ZONE_SPLASH_ZMAX, ":")

    # members, painter's algorithm (far first, by midpoint depth)
    drawn = []
    for mid, m in model["members"].items():
        p1 = np.array(joints[m["j1"]])
        p2 = np.array(joints[m["j2"]])
        cat = cat_by_id.get(mid)
        if cat == "L2 - Component reuse":
            color = fs.CAT_BLUE
        elif cat == "L3 - Downgraded reuse":
            color = fs.CAT_ORANGE
        else:
            color = NOT_IN_SCOPE_COLOR  # excluded (stub/foundation) member
        seg = np.stack([p1, p2])
        sx, sy, depth = _iso_project(seg, center, right, up, forward)
        drawn.append((depth.mean(), sx, sy, color, cat is not None))

    for depth, sx, sy, color, in_scope in sorted(drawn, key=lambda t: t[0]):
        ax.plot(sx, sy, color=color, linewidth=2.2 if in_scope else 1.4,
                 alpha=1.0 if in_scope else 0.6, solid_capstyle="butt", zorder=2)

    ax.set_aspect("equal")
    ax.axis("off")

    handles = [
        Line2D([0], [0], color=fs.CAT_BLUE, linewidth=2.2, label="L2 -- Component reuse"),
        Line2D([0], [0], color=fs.CAT_ORANGE, linewidth=2.2, label="L3 -- Downgraded reuse"),
        Line2D([0], [0], color=NOT_IN_SCOPE_COLOR, linewidth=1.4, alpha=0.6, label="Excluded (stub/foundation)"),
        Line2D([0], [0], color=fs.BASELINE, linewidth=0.7, linestyle=":", label="Splash zone bounds"),
        Line2D([0], [0], color=fs.BASELINE, linewidth=0.7, linestyle=(0, (2, 2)), label="MSL"),
    ]
    fig.legend(handles=handles, loc="lower center", ncol=2, frameon=False, fontsize=7,
               bbox_to_anchor=(0.5, -0.02))
    fig.subplots_adjust(bottom=0.14)

    return fs.save_fig(fig, OUT_DIR, "C3_jacket_reuse_isometric")


# ---------------------------------------------------------------------------
# C4 -- reuse outcome BY MASS, not member count. Reads the live, author-maintained
# embodied_carbon_inputs.xlsx (Calculations sheet, read-only here --
# never regenerated, see docs/decisions.md).
#
# Deliberately a DIFFERENT scope than C1/C3 (which restrict to the 80 "big"
# bay members): this figure's whole point is to show where the excluded
# mass ends up, so it uses the workbook's own scope -- all 104 non-pile
# members (member_id 1-104), i.e. only the below-mudline piles (105-112,
# ~137t of ~2m-diameter pile steel that is never dug up under the assumed
# partial-decommissioning boundary) are left out. Short stub members that
# C1/C3 excluded entirely still appear here -- their full mass legitimately
# flows to L4/Recycling (new_length=0 -> Q col -> S=0, whole member is
# Mass Recycling), which is exactly the "everything we didn't reuse becomes
# recycling" accounting the analysis required.
#
# Mass accounting: per member, `S` (Mass of Reuse Member, new/cut length x
# post-corrosion area) is the reuse-pathway mass; `Z` (Mass Recycling) is
# the cut-off/trim mass (legs only -- for braces the cut-off length is a
# line-model artifact, never real steel, correctly zero, see
# docs/decisions.md). S + Z is the real physical steel
# mass per member; summing S+Z across all 104 members (413.8 t) matches the
# jacket total (674 t) minus the excluded piles (~137t) to within rounding
# -- i.e. nothing is silently lost except that documented brace-overlap
# correction. L1 is included for completeness but is 0 -- no bay currently
# qualifies for structural reuse (matches C1).
# ---------------------------------------------------------------------------

def load_mass_breakdown():
    # Column letters below match the sheet layout as of 25.08.2026 -- the
    # the author inserted an 8-column retrofit-can-thickness lookup (I:P) ahead of
    # the original member_id/category/mass columns, shifting everything that
    # used to start at H to I onward. If this breaks again, re-locate the
    # headers "member_id"/"reuse_level_A"/"Mass of Reuse Member"/"Mass
    # Recycling no thickness increase" directly rather than assuming these
    # letters still hold.
    wb = openpyxl.load_workbook(CARBON_WORKBOOK, data_only=True)
    ws = wb["Calculations"]
    rows = []
    for r in range(4, 108):  # member_id 1-104
        rows.append({
            "member_id": ws[f"I{r}"].value,
            "category": ws[f"X{r}"].value,
            "mass_reuse_kg": ws[f"AB{r}"].value or 0.0,
            "mass_recyc_kg": ws[f"AI{r}"].value or 0.0,
            # retrofit can-thickness steel is added to the recycled cut-off
            # mass (it's thickened at the joint before being cut away), so
            # L4 genuinely differs by scenario -- L2/L3 do not, per
            # docs/decisions.md's "retrofit-INDEPENDENT" note.
            "mass_recyc_A_kg": ws[f"AJ{r}"].value or 0.0,
            "mass_recyc_B_kg": ws[f"AK{r}"].value or 0.0,
        })
    return pd.DataFrame(rows)


def fig_C4_mass_breakdown():
    df = load_mass_breakdown()
    l2_kg = df.loc[df["category"] == "L2", "mass_reuse_kg"].sum()
    l3_kg = df.loc[df["category"] == "L3", "mass_reuse_kg"].sum()
    l4_a_kg = df["mass_recyc_A_kg"].sum()  # cut-off/trim mass incl. Resize A can steel
    l4_b_kg = df["mass_recyc_B_kg"].sum()  # cut-off/trim mass incl. Resize B can steel
    l1_kg = 0.0  # no bay currently qualifies for structural reuse (see C1)
    # Report totals against each retrofit scenario separately -- L4 is the
    # only stage that moves, so the total moves by exactly that much too.
    total_a_kg = l1_kg + l2_kg + l3_kg + l4_a_kg
    total_b_kg = l1_kg + l2_kg + l3_kg + l4_b_kg

    stages = [
        ("L1\nStructural reuse", l1_kg, None, fs.BASELINE),
        ("L2\nComponent reuse", l2_kg, None, fs.CAT_BLUE),
        ("L3\nDowngraded reuse", l3_kg, None, fs.CAT_ORANGE),
        ("L4\nRecycle", l4_a_kg, l4_b_kg, fs.CAT_AQUA),
    ]

    fig, ax = plt.subplots(figsize=fs.usable_figsize(width_frac=0.75, aspect=0.6))
    x = np.arange(len(stages))
    labels = [s[0] for s in stages]
    heights_t = [s[1] / 1000 for s in stages]
    upper_t = [(s[2] / 1000 if s[2] is not None else s[1] / 1000) for s in stages]
    max_h = max(upper_t)

    for xi, (label, lo_kg, hi_kg, color) in zip(x, stages):
        lo_t = lo_kg / 1000
        if hi_kg is None:
            ax.bar(xi, lo_t, color=color, width=0.55, zorder=3,
                   edgecolor=fs.MARKER_EDGE, linewidth=fs.MARKER_EDGE_WIDTH)
            ax.text(xi, lo_t + max_h * 0.03, f"{lo_t:.1f} t",
                     ha="center", va="bottom", fontsize=7.5, color=fs.INK_PRIMARY)
        else:
            hi_t = hi_kg / 1000
            # Solid segment = Resize A value; hatched upper segment = the
            # extra mass under Resize B -- one bar, range read top-to-bottom.
            ax.bar(xi, lo_t, color=color, width=0.55, zorder=3,
                   edgecolor=fs.MARKER_EDGE, linewidth=fs.MARKER_EDGE_WIDTH)
            ax.bar(xi, hi_t - lo_t, bottom=lo_t, color=color, alpha=0.35,
                   hatch="////", width=0.55, zorder=3,
                   edgecolor=fs.MARKER_EDGE, linewidth=fs.MARKER_EDGE_WIDTH)
            ax.text(xi, hi_t + max_h * 0.075, f"{lo_t:.1f}–{hi_t:.1f} t",
                     ha="center", va="bottom", fontsize=7.5, color=fs.INK_PRIMARY)
            ax.text(xi, hi_t + max_h * 0.03, "Resize A–B", ha="center",
                     va="bottom", fontsize=6, color=fs.INK_SECONDARY, style="italic")

    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=7.5, color=fs.INK_PRIMARY)
    ax.set_ylabel("Mass (t)")
    ax.set_ylim(0, max_h * 1.35)
    ax.spines[["top", "right"]].set_visible(False)
    ax.spines["left"].set_color(fs.BASELINE)
    ax.spines["bottom"].set_color(fs.BASELINE)
    ax.tick_params(colors=fs.INK_SECONDARY)
    ax.grid(axis="y", color=fs.GRIDLINE, linewidth=0.6, zorder=0)
    ax.set_axisbelow(True)

    fig.tight_layout()
    return fs.save_fig(fig, OUT_DIR, "C4_mass_breakdown"), (total_a_kg, total_b_kg)


# ---------------------------------------------------------------------------
# C5 -- avoided emissions (tCO2e), NOT mass. Two baselines (what the L2/L3
# reuse credit is measured against: BOF or EAF new-steel production) on the
# x-axis, each a stacked bar of Reuse (L2+L3, fixed, retrofit-independent)
# and Recycle (L4, retrofit-dependent -- same Resize A/B range convention as
# C4). Reads EF_BOD/EF_EAF/EF_REUSE and the masses straight from the live
# Calculations sheet -- never hardcode the EF values here, they're the
# the author's own live inputs (D13/D14/D15).
#
# NOTE: L3's reuse credit is computed here as (EF_BOD-EF_REUSE) directly
# from the masses, NOT read from the sheet's own C22 cell -- C22 currently
# has a formula bug (subtracts EF_EAF instead of EF_REUSE, unlike C21/E22
# which are both correct). Recompute from D13/D14/D15 + Q3.mass_reuse_kg
# once C22 is fixed in the sheet; this function does not depend on it.
# ---------------------------------------------------------------------------

def fig_C5_carbon_savings_by_baseline():
    wb = openpyxl.load_workbook(CARBON_WORKBOOK, data_only=True)
    ws = wb["Calculations"]
    ef_bod = ws["D13"].value
    ef_eaf = ws["D14"].value
    ef_reuse = ws["D15"].value

    df = load_mass_breakdown()
    reuse_mass_kg = df.loc[df["category"].isin(["L2", "L3"]), "mass_reuse_kg"].sum()
    l4_a_kg = df["mass_recyc_A_kg"].sum()
    l4_b_kg = df["mass_recyc_B_kg"].sum()

    reuse_bof_t = (ef_bod - ef_reuse) * reuse_mass_kg / 1000
    reuse_eaf_t = (ef_eaf - ef_reuse) * reuse_mass_kg / 1000
    recyc_a_t = (ef_bod - ef_eaf) * l4_a_kg / 1000
    recyc_b_t = (ef_bod - ef_eaf) * l4_b_kg / 1000

    baselines = [
        ("vs BOF baseline", reuse_bof_t),
        ("vs EAF baseline", reuse_eaf_t),
    ]

    fig, ax = plt.subplots(figsize=fs.usable_figsize(width_frac=0.75, aspect=0.6))
    x = np.arange(len(baselines))
    width = 0.5

    reuse_heights = [b[1] for b in baselines]
    ax.bar(x, reuse_heights, width=width, color=fs.CAT_BLUE, zorder=3,
           edgecolor=fs.MARKER_EDGE, linewidth=fs.MARKER_EDGE_WIDTH,
           label="Reuse (L2+L3)")
    ax.bar(x, [recyc_a_t] * len(x), width=width, bottom=reuse_heights,
           color=fs.CAT_AQUA, zorder=3, edgecolor=fs.MARKER_EDGE,
           linewidth=fs.MARKER_EDGE_WIDTH, label="Recycle (L4, Resize A)")
    ax.bar(x, [recyc_b_t - recyc_a_t] * len(x), width=width,
           bottom=[r + recyc_a_t for r in reuse_heights], color=fs.CAT_AQUA,
           alpha=0.35, hatch="////", zorder=3, edgecolor=fs.MARKER_EDGE,
           linewidth=fs.MARKER_EDGE_WIDTH, label="+ Resize B")

    total_lo = [r + recyc_a_t for r in reuse_heights]
    total_hi = [r + recyc_b_t for r in reuse_heights]
    max_h = max(total_hi)
    for xi, r_t, lo_t, hi_t in zip(x, reuse_heights, total_lo, total_hi):
        ax.text(xi, hi_t + max_h * 0.03, f"{lo_t:.0f}–{hi_t:.0f} tCO₂e",
                 ha="center", va="bottom", fontsize=7.5, color=fs.INK_PRIMARY)
        ax.text(xi, r_t / 2, f"{r_t:.0f}", ha="center", va="center",
                 fontsize=7.5, color="white")
        ax.text(xi, r_t + recyc_a_t / 2, f"{recyc_a_t:.0f}–{recyc_b_t:.0f}",
                 ha="center", va="center", fontsize=7, color="white")

    ax.set_xticks(x)
    ax.set_xticklabels([b[0] for b in baselines], fontsize=8, color=fs.INK_PRIMARY)
    ax.set_ylabel("Avoided emissions (tCO₂e)")
    ax.set_ylim(0, max_h * 1.2)
    ax.spines[["top", "right"]].set_visible(False)
    ax.spines["left"].set_color(fs.BASELINE)
    ax.spines["bottom"].set_color(fs.BASELINE)
    ax.tick_params(colors=fs.INK_SECONDARY)
    ax.grid(axis="y", color=fs.GRIDLINE, linewidth=0.6, zorder=0)
    ax.set_axisbelow(True)
    ax.legend(loc="upper right", fontsize=7, frameon=False)

    fig.tight_layout()
    return fs.save_fig(fig, OUT_DIR, "C5_carbon_savings_by_baseline")


# ---------------------------------------------------------------------------
# Tables -- plain markdown, for pasting into the thesis (matches the
# Table 13/14 workflow in docs/decisions.md).
# ---------------------------------------------------------------------------

def table_member_summary(members):
    rows = []
    for zone in ["atmospheric", "submerged", "splash"]:
        sub = members[members["zone"] == zone]
        n = len(sub)
        n_l2 = int((sub["reuse_category_A"] == "L2 - Component reuse").sum())
        n_l3 = int((sub["reuse_category_A"] == "L3 - Downgraded reuse").sum())
        rows.append((zone, n, n_l2, n_l3))
    df = pd.DataFrame(rows, columns=["Zone", "n members", "L2 -- Component reuse",
                                       "L3 -- Downgraded reuse"])
    total = pd.DataFrame([["All", len(members),
                            int((members["reuse_category_A"] == "L2 - Component reuse").sum()),
                            int((members["reuse_category_A"] == "L3 - Downgraded reuse").sum())]],
                          columns=df.columns)
    df = pd.concat([df, total], ignore_index=True)
    return df


def table_bay_summary(bays):
    df = bays[["bay_id", "n_members", "n_joints", "worst_joint_D_A_K",
               "worst_joint_D_B_K", "structural_reuse_pass_A",
               "structural_reuse_pass_B"]].copy()
    df.columns = ["Bay", "n members", "n joints", "Worst D (Retrofit A)",
                  "Worst D (Retrofit B)", "L1 pass -- A", "L1 pass -- B"]
    df["Bay"] = df["Bay"].apply(lambda b: f"Bay {b}")
    for c in ["Worst D (Retrofit A)", "Worst D (Retrofit B)"]:
        df[c] = df[c].map(lambda v: f"{v:.2f}")
    return df


def df_to_markdown(df):
    cols = [str(c) for c in df.columns]
    rows = [[str(v) for v in row] for row in df.itertuples(index=False)]
    widths = [max(len(c), *(len(r[i]) for r in rows)) if rows else len(c)
              for i, c in enumerate(cols)]
    def fmt_row(vals):
        return "| " + " | ".join(v.ljust(w) for v, w in zip(vals, widths)) + " |"
    lines = [fmt_row(cols), "| " + " | ".join("-" * w for w in widths) + " |"]
    lines += [fmt_row(r) for r in rows]
    return "\n".join(lines)


if __name__ == "__main__":
    fs.apply_style()
    members, bays = load()

    fig_C1_cascade(members, bays)
    fig_C2_bay_margin_dumbbell(bays)
    fig_C3_jacket_reuse_isometric(members)
    (_, (total_a_kg, total_b_kg)) = fig_C4_mass_breakdown()
    print(f"C4: total mass accounted = {total_a_kg/1000:.2f}-{total_b_kg/1000:.2f} t "
          f"(Resize A-B, 104 non-pile members)")

    t1 = table_member_summary(members)
    t2 = table_bay_summary(bays)

    out_txt = OUT_DIR / "reuse_tables.md"
    out_txt.parent.mkdir(parents=True, exist_ok=True)
    with open(out_txt, "w", encoding="utf-8") as f:
        f.write("### Table -- member-level reuse outcome by zone\n\n")
        f.write(df_to_markdown(t1))
        f.write("\n\n### Table -- bay-level structural-reuse (L1) diagnostic\n\n")
        f.write(df_to_markdown(t2))
        f.write("\n")

    print(f"Saved figures to {OUT_DIR}")
    print(f"Saved tables to {out_txt}")
    print()
    print(df_to_markdown(t1))
    print()
    print(df_to_markdown(t2))
